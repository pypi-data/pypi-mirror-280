import argparse
import os
import sys
from typing import List

import pymongo
from openpyxl import load_workbook,workbook, worksheet

class Array2d:

    def __init__(self, tuples:List):
        self._array = []

        compare = None
        for i,t in enumerate(tuples):
            self._array.append(t)
            if compare is None:
                compare = t
            elif len(compare) != len(t) :
                raise ValueError("List of tuples to ctor are not uniform (lengths differ")

    def xy(self,  x, y):
        return self.array[y][x]

    def len_x(self):
        return len(self._array[0])

    def len_y(self):
        return len(self._array)

    def row_first(self):
        for row in self._array:
            for col in row:
                yield col

    def col_first(self):
        for x in range(len(self._array)):
            for y in range(len(self._array[0])):
                yield self._array[y][x]

    def row_order(self):
        for row in self._array:
            yield row

    def col_order(self):

        for x in range(len(self._array[0])):
            col = []
            for y in range(len(self._array)):
                col.append(self._array[y][x])
            yield col

    def __repr__(self):
        s="[\n"
        for row in self._array:
            for col in row:
                s = f"{s} {col:5},"
            s = f"{s}\n"
        s=f"{s}]\n"
        return s

class ExcelWorkbook:

    def __init__(self, excel_filename):
        self._workbook = load_workbook(filename=excel_filename, read_only=True)
        self._sheets = {}
        for name in self._workbook.sheetnames:
            self._sheets[name] = self._workbook[name]

    @property
    def workbook(self):
        return self._workbook

    @property
    def sheet_names(self):
        return list(self._sheets.keys())

    @property
    def sheets(self):
        return list(self._sheets.itervalues())

    def sheet(self, name):
        return self._sheets[name]


class ExcelSheet:

    def __init__(self, workbook, sheet_name="None"):

        self._workbook = workbook
        self._sheet_name = sheet_name
        if sheet_name:
            self._sheet = self._workbook[self._sheet_name]
        else:
            self._sheet = self._workbook.active

    def sheet_name(self):
        return self._sheet.filename

    @property
    def sheet(self):
        return self._sheet

if __name__ == "__main__":

    parser = argparse.ArgumentParser()

    parser.add_argument("--host", default="mongodb://localhost:27017")
    parser.add_argument("--excelfile")
    parser.add_argument("--sheetname")
    parser.add_argument("--minrow", type=int)
    parser.add_argument("--mincol", type=int)
    parser.add_argument("--maxrow", type=int)
    parser.add_argument("--maxcol", type=int)
    parser.add_argument("--database", default="census")
    parser.add_argument("--collection", default="survey")
    parser.add_argument("--lowerright")
    parser.add_argument("--drop", action="store_true", default=False)
    parser.add_argument("--colorder", action="store_true", default=False)
    args = parser.parse_args()

    client = pymongo.MongoClient(args.host)
    db = client[args.database]
    collection = db[args.collection]

    if args.drop:
        print(f"Dropping collection {args.collection}")
        db.drop_collection(args.collection)

    print(f"Writing data to cluster {args.host} collection:{args.database}.{args.collection}")

    if args.excelfile:
        if os.path.isfile(args.excelfile):
            cb = ExcelWorkbook(args.excelfile)
        else:
            print(f"{args.excelfile} is not a file")
            sys.exit(1)
    else:
        print(f'--excelfile is not specified')
        sys.exit(1)

    if args.sheetname:
        sh = ExcelSheet(cb.workbook, args.sheetname)
    else:
        print(f'--sheetname is not specified')
        sys.exit(1)



    doc = {}

    count = 1

    col_index = {}
    count = 1

    #
    # Get the column keys
    print("Row keys")
    for row in sh.sheet.iter_rows(min_row=args.minrow,
                                   min_col=args.mincol,
                                   max_row=args.minrow,
                                   max_col=args.maxcol,
                                   values_only=True):

        for v in row:
            if v is None:
                continue
            else:

                col_index[count] = v
                print(f"{count}. {v}", end=" ")
                count = count + 1
        print("")

    row_index = {}
    count = 1

    #
    # get the row keys
    print("Column keys")
    for row in sh.sheet.iter_rows(min_row=args.minrow,
                                  min_col=args.mincol,
                                  max_row=args.maxrow,
                                  max_col=args.mincol,
                                  values_only=True):
        for v in row:
            if v is None:
                continue
            else:
                row_index[count] = v
                print(f"{count}. {v}")
                count = count + 1
    print("")
    row_count = 1
    col_count = 1

    doc_count = 0

    array2d = []
    index = 1
    for row in sh.sheet.iter_rows(min_row=args.minrow+1,
                                  min_col=args.mincol+1,
                                  max_row=args.maxrow,
                                  max_col=args.maxcol,
                                  values_only=True):
        array2d.append(row)



    # print(array2d[0][0])
    # y = len(array2d)
    # x= len(array2d[0])
    # print(f"x={x} y={y}")
    # print(f"x=3,y=2 {array2d[2][3]}")
    a2d = Array2d(array2d)
    # print(a2d)
    # print("row_order")
    # for i in a2d.row_order():
    #     print(i)
    #
    # print("col_order")
    # for i in a2d.col_order():
    #     print(i)
    #
    # for i in a2d.row_first():
    #     print(f"{i:5}", end=" ")
    # print("")
    #
    # for i in a2d.col_first():
    #     print(f"{i:5}", end=" ")
    # print("")


    row_count = 1
    col_count = 1
    if args.colorder:
        for col in a2d.col_order():
            doc = {}
            doc_count = doc_count + 1
            doc["title"] = col_index[col_count]
            col_count = col_count + 1
            row_count = 1
            for v in col:
                doc[row_index[row_count]] = v
                row_count = row_count + 1
                #print(f"colorder {doc}")
            print(f"{doc_count} {doc}")
            collection.insert_one(doc)
    else:
        for row in a2d.row_order():
            doc = {}
            doc_count = doc_count + 1
            doc["title"] = row_index[row_count]
            row_count = row_count + 1
            col_count = 1
            for v in row:
                doc[col_index[col_count]] = v
                col_count = col_count + 1
            print(f"{doc_count} {doc}")
            collection.insert_one(doc)


    # for row in sh.sheet.iter_rows(min_row=args.minrow+1,
    #                               min_col=args.mincol+1,
    #                               max_row=args.maxrow,
    #                               max_col=args.maxcol,
    #                               values_only=True):
    #     doc={}
    #     doc_count = doc_count + 1
    #     if args.colorder:
    #         doc["title"] = col_index[col_count]
    #         col_count = col_count + 1
    #         row_count = 1
    #         for v in row:
    #             doc[row_index[row_count]] = v
    #             row_count = row_count + 1
    #             #print(f"colorder {doc}")
    #     else:
    #         doc["title"] = row_index[row_count]
    #         row_count = row_count + 1
    #         col_count = 1
    #         for v in row:
    #             doc[col_index[col_count]] = v
    #             col_count = col_count + 1
    #
    #
    #     print(f"{doc_count} {doc}")
    #     collection.insert_one(doc)






