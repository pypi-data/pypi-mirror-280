import pprint

from openpyxl import load_workbook,workbook, worksheet
wb = load_workbook('emeadevit.xlsx')
#print(wb.sheetnames)


question_doc = {
    "Total" : 1516,
    "row" : 12,
    "column": 3,
    "Gender": {
        "row" : 12,
        "column": 4,
        "Male" : 1015,
        "Female": 501,
    },
    "Age" : {
        "row": 12,
        "column":4,
        "16-24" : 84,
        "25-34" : 606,
        "35-44" : 610,
        "45-54" : 144,
        "55+"   : 72
    },
    "Country" : {
        "row": 12,
        "column": 11,
        "Germany" : 513,
        "UK" : 502,
        "France" : 501,
    },
    "Industry Sector" : {
        "row": 12,
        "column": 14,
        "Architecture, Engineering & Building" : 72,
        "Arts & Culture" : 10,
        "Education" : 15,
        "Finance" : 77,
        "Healthcare": 38,
        "HR" : 13,
        "IT & Telecoms" : 1021,
        "Legal" : 5,
        "Manufacturing and Utilities" : 113,
        "Retail, Catering and Leisure": 71,
        "Sales, Media & Marketing" : 17,
        "Travel & Transport": 27,
        "Other": 42
    },
    "Company Size" : {
        "Sole Trader" : 8,
        "1-9 employees": 30,
        "10-49 employees": 61,
        "50-99 employees": 116,
        "100-249 employees": 116,
        "250-500 employees": 382,
        "More than 500 employees": 666,
    },
    "IT Decision Maker vs Developers" : {
        "IT Decision Maker" :756,
        "Developers": 760
    }
}

# for filename in wb.sheetnames:
#     ws = wb[filename]
#     question = ws["A10"].value
#     print(f"{question}")


def get_responses(ws:worksheet, row=16):

    responses = {}
    while True:
        response = ws.cell(row=row, column=2).value
        if response is None:
            break
        else:
            responses[response] = None
        row = row + 2
    return responses


def get_last_column(template:dict):
    column = 0
    for key, value in template.items():
        if type(value) is dict:
            column = column + get_last_column(value)
        elif key == "row":
            continue
        elif key == "column":
            continue
        else:
            column = column + 1
    return column


def get_response_values(ws:worksheet, template:dict, responses, init_row:int, init_column:int):
    """

    :param ws: Worksheet
    :param template: example doc we expect to get, gives us the keys
    :param responses: the document of values we are constructing
    :param init_row: Which row to start reading responses from
    :param init_column: Which column to start reading responses fromn
    :return:
    """
    response_values = {}
    width = get_last_column(template)

    max_response_rows = (len(responses) * 2) + init_row - 1
    for col in ws.iter_cols(min_row=init_row, max_row=max_response_rows, min_col=init_column, max_col=init_column+width-1, values_only=True):
        for k,v in template.items():
            response_values[k] = None
            if k == "row" or k == "column":
                continue
            elif type(v) is dict:
                for nk in v.keys():
                    if k == "row" or k == "column":
                        continue
                    else:
                        response_values[k][nk] = col[template[k][nk]["column"]]
            else:
                response_values[k] = col[template["column"]]
    print(response_values)

def get_questions(ws:worksheet, row=16):

    question = {}

    while True:
        main_question = ws.cell( row=row, column=1).value
        if main_question is None:
            break
        else:
            question["question"] = {main_question:ws.cell(row=10, column=1).value}
            responses = get_responses(ws, row=row)
            question["responses"]=responses
            pprint.pprint(question)

            # k=input("Next..")
        row = row + len(responses) * 2 + 1

    get_response_values(ws=ws, template=question_doc, responses=responses, init_row=16, init_column=3)

for name in wb.sheetnames:
    ws = wb[name]
    #print( ws.cell(row=16, column=2).value)
    print(f"Questions in sheet '{ws.title}'")
    get_questions(ws)
