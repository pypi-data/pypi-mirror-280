import warnings

try:
    import openpyxl
    import webcolors
    import pandas as pd
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
except ImportError:
    raise 'pip install ninjatools[excel] or ninjatools[all] to use excel functions!'

# Remove warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class ColorObject(object):
    def __init__(self, hex_, rgb, color_name):
        self.hex = hex_
        self.rgb = rgb
        self.name = color_name

    def __repr__(self):
        str_color = ', '.join(str(_) for _ in self.rgb)
        return f'Closest Color: {self.name} | RGB: ({str_color})'


class Excel:
    def __init__(self, workbook_path):
        self.wb_path = workbook_path
        self.sheet_name = None
        self.wb = None

    def load_workbook(self):
        self.wb = openpyxl.load_workbook(self.wb_path, data_only=True)

    def get_sheets(self) -> list:
        """
        Returns a list of sheet names
        :return:
        """
        return [_ for _ in self.wb.sheetnames]

    # Get cell_value using openpyxl
    def cell(self, cell, sheet_name=None):
        """
        Returns the value of a cell
        :param cell:
        :param sheet_name:
        :return:
        """

        if isinstance(cell, str):
            cell = self.get_cell(cell)
        elif not isinstance(cell, tuple) and not isinstance(cell, list):
            raise TypeError('cell must be a tuple or string')

        sheet_name = sheet_name if sheet_name else self.sheet_name
        ws = self.wb[sheet_name]
        value = ws.cell(cell[0], cell[1]).value
        value = '' if value is None else value
        value = str(value).strip().replace('\xa0', ' ')
        return value

    def get_color(self, cell) -> ColorObject:
        """
        Returns the color of a cell
        :param cell:
        :return:
        """
        ws = self.wb[self.sheet_name]
        hex_color = ws[cell].fill.start_color.index
        hex_color = f'#{hex_color[2:]}'
        rgb_color = webcolors.hex_to_rgb(hex_color)
        rgb_color = (rgb_color.red, rgb_color.green, rgb_color.blue)
        color_name = self.closest_color(rgb_color)

        return ColorObject(hex_color, rgb_color, color_name)

    def read_range(self, start, end, sheet_name=None, remove_empty=False) -> list:
        """
        Reads a range of cells
        :param start:
        :param end:
        :param sheet_name:
        :param remove_empty:
        :return:
        """
        data = []

        sheet_name = sheet_name if sheet_name else self.sheet_name

        rc_start = self.get_cell(start)
        rc_end = self.get_cell(end)

        for row in range(rc_start[0], rc_end[0] + 1):
            temp = []
            for col in range(rc_start[1], rc_end[1] + 1):
                temp.append(self.cell((row, col), sheet_name))

            if remove_empty:
                if not all([not _ for _ in temp]):
                    data.append(temp)
            else:
                data.append(temp)

        return data

    def read_all(self, sheet_name=None, remove_empty=False) -> list:
        """
        Reads all cells in a sheet
        :param sheet_name:
        :param remove_empty:
        :return:
        """
        # Read excel
        df = pd.read_excel(self.wb_path, sheet_name=self.sheet_name)

        # Remove all NaN columns
        df.dropna(axis=1, how='all', inplace=True)

        # Convert DataFrame to list of lists
        data = df.values.tolist()

        # Add headers at the start
        data.insert(0, df.columns.tolist())
        return data

    # Still buggy
    def read_indefinitely(self, start_cell: str, num_of_columns: int, num_of_rows=None, steps: int = 1) -> list:
        """
        Reads a range of cells indefinitely
        :param start_cell:
        :param num_of_columns:
        :param num_of_rows:
        :param steps:
        :return:
        """

        data = []

        ord_num = ord(start_cell[0])
        i = int(start_cell[1:])

        while True:
            try:
                data.append([self.cell(f'{chr(ord_num + _)}{i}') for _ in range(num_of_columns)])
                if num_of_rows:
                    if len(data) >= num_of_rows:
                        break
            except (IndexError, AttributeError):
                if not data:
                    print(f'Check if the cell {start_cell} to # of columns is not all empty!')
                break
            except (Exception,):
                break
            i += steps

        return data

    def get_last(self) -> str:
        """
        Returns the last cells
        :return:
        """
        ws = self.wb[self.sheet_name]
        return self.get_cell_name(ws.max_row, ws.max_column)

    @staticmethod
    def get_cell(cell) -> tuple:
        """
        Returns the cell coordinates
        :param cell:
        :return:
        """
        xy = coordinate_from_string(cell)
        col = column_index_from_string(xy[0])
        row = xy[1]
        return row, col

    @staticmethod
    def get_cell_name(row, column) -> str:
        """
        Returns the cell name
        :param row:
        :param column:
        :return:
        """
        return openpyxl.utils.get_column_letter(column) + str(row)

    @staticmethod
    def convert_to_dict(data: list, key_idx: int, header_idx: int = 0) -> dict:
        """
        Converts a list to a dictionary
        :param data:
        :param key_idx:
        :param header_idx:
        :return:
        """
        dict_data = {}
        for idx, row in enumerate(data):
            if idx > 0:
                dict_data[row[key_idx]] = dict(zip(data[header_idx], row))

        return dict_data

    @staticmethod
    def closest_color(rgb):
        differences = {}
        for color_hex, color_name in webcolors.CSS3_HEX_TO_NAMES.items():
            r, g, b = webcolors.hex_to_rgb(color_hex)
            differences[sum([(r - rgb[0]) ** 2,
                             (g - rgb[1]) ** 2,
                             (b - rgb[2]) ** 2])] = color_name
        return differences[min(differences.keys())]

    # TODO: Cell write values/formulas
